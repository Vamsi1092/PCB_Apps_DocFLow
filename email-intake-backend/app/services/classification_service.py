import fitz
from openpyxl import load_workbook
import os


def classify_email(
    subject,
    body_preview,
    attachment_filename,
    file_type,
    document_content=""
):

    subject = subject or ""
    body_preview = body_preview or ""
    attachment_filename = attachment_filename or ""
    file_type = file_type or ""
    document_content = document_content or ""

    combined_text = (
    subject + " " +
    body_preview + " " +
    attachment_filename + " " +
    file_type + " " +
    document_content
    ).lower()

    if (
        "invoice" in combined_text
        or " inv" in combined_text
        or "bill" in combined_text
    ):
        return {
            "classification": "Invoice",
            "confidence": 0.90,
            "reason": "Invoice keyword detected",
            "target_folder": "Invoice"
        }

    if (
        "purchase order" in combined_text
        or " po " in combined_text
        or "order" in combined_text
    ):
        return {
            "classification": "Purchase Order",
            "confidence": 0.90,
            "reason": "Purchase Order keyword detected",
            "target_folder": "Purchase Order"
        }

    if (
        "acknowledgement" in combined_text
        or "acknowledgment" in combined_text
        or " ack " in combined_text
    ):
        return {
            "classification": "Acknowledgement",
            "confidence": 0.90,
            "reason": "Acknowledgement keyword detected",
            "target_folder": "Acknowledgement"
        }

    return {
        "classification": "Others",
        "confidence": 0.50,
        "reason": "No matching business keywords found",
        "target_folder": "Others"
    }

def read_excel_content(file_path):

    try:

        workbook = load_workbook(file_path)

        sheet = workbook.active

        content = []

        for row in sheet.iter_rows(values_only=True):

            row_text = " | ".join(
                [str(cell) for cell in row if cell is not None]
            )

            if row_text:
                content.append(row_text)

        return "\n".join(content[:20])

    except Exception as e:

        return f"Error reading file: {str(e)}"

from docx import Document

def read_docx_content(file_path):

    document = Document(file_path)

    text = []

    for paragraph in document.paragraphs:
        if paragraph.text.strip():
            text.append(paragraph.text.strip())

    return "\n".join(text)

def read_pdf_content(file_path):

    try:
        document = fitz.open(file_path)

        text = []

        for page in document:
            page_text = page.get_text()
            if page_text.strip():
                text.append(page_text.strip())

        document.close()

        return "\n".join(text)

    except Exception as e:
        return f"Error reading PDF file: {str(e)}"